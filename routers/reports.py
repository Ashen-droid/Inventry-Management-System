"""Reports router: PDF, Excel, CSV exports for Sales/Purchase/Profit/Stock."""
import io, csv
import datetime
# pyrefly: ignore [missing-import]
from fastapi import APIRouter, Depends, Query
# pyrefly: ignore [missing-import]
from fastapi.responses import StreamingResponse
# pyrefly: ignore [missing-import]
from sqlalchemy.orm import Session
# pyrefly: ignore [missing-import]
from sqlalchemy import func

from database import (get_db, Sale, SaleItem, Purchase, PurchaseItem,
                      Product, Category, User)
from services.auth_service import get_current_user
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/reports", tags=["Reports"])

HEADER_COLOR = colors.HexColor("#1a237e")
ROW_ALT      = colors.HexColor("#f5f5f5")


def _date_filter(start: str, end: str):
    start_dt = datetime.datetime.strptime(start, "%Y-%m-%d")
    end_dt   = datetime.datetime.strptime(end,   "%Y-%m-%d") + datetime.timedelta(days=1)
    return start_dt, end_dt


# ────────────────────────────────────────────────────────
#  SALES REPORT
# ────────────────────────────────────────────────────────
@router.get("/sales")
def sales_report(
    format: str = Query("pdf", enum=["pdf", "excel", "csv"]),
    start:  str = Query(datetime.date.today().replace(day=1).isoformat()),
    end:    str = Query(datetime.date.today().isoformat()),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start_dt, end_dt = _date_filter(start, end)
    sales = (db.query(Sale)
             .filter(Sale.sale_date >= start_dt, Sale.sale_date < end_dt)
             .order_by(Sale.sale_date.desc()).all())

    rows = [["Invoice #", "Date", "Customer", "Items", "Total (Rs.)"]]
    for s in sales:
        rows.append([
            s.invoice_number,
            s.sale_date.strftime("%Y-%m-%d %H:%M"),
            s.customer.name if s.customer else "Walk-in",
            len(s.items),
            f"{s.total_amount:,.2f}",
        ])
    total_row = ["", "", "", "TOTAL", f"{sum(s.total_amount for s in sales):,.2f}"]
    rows.append(total_row)

    title = f"Sales Report ({start} to {end})"
    return _generate_report(rows, title, format, "sales_report")


# ────────────────────────────────────────────────────────
#  PURCHASE REPORT
# ────────────────────────────────────────────────────────
@router.get("/purchases")
def purchase_report(
    format: str = Query("pdf", enum=["pdf", "excel", "csv"]),
    start:  str = Query(datetime.date.today().replace(day=1).isoformat()),
    end:    str = Query(datetime.date.today().isoformat()),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start_dt, end_dt = _date_filter(start, end)
    purchases = (db.query(Purchase)
                 .filter(Purchase.purchase_date >= start_dt, Purchase.purchase_date < end_dt)
                 .order_by(Purchase.purchase_date.desc()).all())

    rows = [["Purchase #", "Date", "Supplier", "Items", "Total Cost (Rs.)"]]
    for p in purchases:
        rows.append([
            p.purchase_id,
            p.purchase_date.strftime("%Y-%m-%d"),
            p.supplier.name if p.supplier else "-",
            len(p.items),
            f"{p.total_cost:,.2f}",
        ])
    rows.append(["", "", "", "TOTAL", f"{sum(p.total_cost for p in purchases):,.2f}"])

    return _generate_report(rows, f"Purchase Report ({start} to {end})", format, "purchase_report")


# ────────────────────────────────────────────────────────
#  STOCK REPORT
# ────────────────────────────────────────────────────────
@router.get("/stock")
def stock_report(
    format: str = Query("pdf", enum=["pdf", "excel", "csv"]),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    products = db.query(Product).filter(Product.is_active == True).all()

    rows = [["Product", "Category", "Price (Rs.)", "Cost Price (Rs.)", "Stock", "Status"]]
    for p in products:
        status = "OUT OF STOCK" if p.quantity == 0 else ("LOW STOCK" if p.quantity <= p.low_stock_threshold else "OK")
        rows.append([
            p.name,
            p.category.name if p.category else "-",
            f"{p.price:,.2f}",
            f"{p.cost_price:,.2f}" if p.cost_price else "-",
            p.quantity,
            status,
        ])

    return _generate_report(rows, "Stock Report", format, "stock_report")


# ────────────────────────────────────────────────────────
#  PROFIT REPORT
# ────────────────────────────────────────────────────────
@router.get("/profit")
def profit_report(
    format: str = Query("pdf", enum=["pdf", "excel", "csv"]),
    start:  str = Query(datetime.date.today().replace(day=1).isoformat()),
    end:    str = Query(datetime.date.today().isoformat()),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    start_dt, end_dt = _date_filter(start, end)

    revenue = db.query(func.sum(Sale.total_amount)).filter(
        Sale.sale_date >= start_dt, Sale.sale_date < end_dt).scalar() or 0
    cost    = db.query(func.sum(Purchase.total_cost)).filter(
        Purchase.purchase_date >= start_dt, Purchase.purchase_date < end_dt).scalar() or 0

    rows = [
        ["Metric", "Amount (Rs.)"],
        ["Total Revenue",  f"{revenue:,.2f}"],
        ["Total Cost",     f"{cost:,.2f}"],
        ["Gross Profit",   f"{revenue - cost:,.2f}"],
        ["Profit Margin",  f"{((revenue-cost)/revenue*100):.1f}%" if revenue > 0 else "N/A"],
    ]
    return _generate_report(rows, f"Profit Report ({start} to {end})", format, "profit_report")


# ────────────────────────────────────────────────────────
#  Helper: generate in chosen format
# ────────────────────────────────────────────────────────
def _generate_report(rows: list, title: str, format: str, filename_base: str):
    if format == "pdf":
        return _to_pdf(rows, title, filename_base)
    elif format == "excel":
        return _to_excel(rows, title, filename_base)
    else:
        return _to_csv(rows, filename_base)


def _to_pdf(rows, title, filename_base):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = [Paragraph(title, styles["Title"]), Spacer(1, 0.5*cm)]

    table = Table(rows, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), HEADER_COLOR),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, ROW_ALT]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#e0e0e0")),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
    ])
    table.setStyle(style)
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename_base}.pdf"})


def _to_excel(rows, title, filename_base):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1a237e")
    alt_fill     = PatternFill("solid", fgColor="F5F5F5")

    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=r_idx+1, column=c_idx+1, value=str(val))
            if r_idx == 0:
                cell.font = header_font
                cell.fill = header_fill
            elif r_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"})


def _to_csv(rows, filename_base):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows([[str(c) for c in row] for row in rows])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"})
